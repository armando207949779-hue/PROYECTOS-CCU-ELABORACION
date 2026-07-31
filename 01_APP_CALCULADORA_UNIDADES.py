# =====================================================
# 01_APP_CALCULADORA_UNIDADES.py
# VERSION 8
# CALCULADORA Y REGISTRO DE TRAZABILIDAD
# PROYECTO CCU - ELABORACIÓN
# =====================================================

import base64
import hashlib
import json
import re
import uuid
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import pandas as pd
import streamlit as st


# =====================================================
# CONFIGURACIÓN GENERAL
# =====================================================

st.set_page_config(
    page_title="Calculadora de Unidades - CCU",
    page_icon="🧮",
    layout="wide",
    initial_sidebar_state="collapsed"
)


# =====================================================
# RUTAS
# =====================================================

RUTA_BASE = Path(__file__).parent

ARCHIVO_BBDD = RUTA_BASE / "BBDD_UNIDADES_MATERIAS_PRIMAS.parquet"
ARCHIVO_LOGO = RUTA_BASE / "CCU_LOGO.png"


# =====================================================
# ESTILOS
# =====================================================

st.markdown(
    """
    <style>
    .block-container {
        padding-top: 1rem;
        padding-bottom: 2rem;
        max-width: 1180px;
    }

    header[data-testid="stHeader"] {
        background: rgba(0,0,0,0);
        height: 0rem;
    }

    div[data-testid="stToolbar"],
    div[data-testid="stDecoration"],
    #MainMenu,
    footer {
        visibility: hidden;
        height: 0%;
    }

    .logo-container {
        display: flex;
        justify-content: center;
        align-items: center;
        margin-top: 0px;
        margin-bottom: 8px;
        width: 100%;
    }

    .logo-container img {
        max-width: 112px;
        height: auto;
        display: block;
    }

    .titulo-principal {
        text-align: center;
        font-size: 30px;
        font-weight: 800;
        color: #003865;
        margin-top: 0px;
        margin-bottom: 2px;
        line-height: 1.2;
    }

    .subtitulo {
        text-align: center;
        font-size: 15px;
        color: #666666;
        margin-bottom: 22px;
    }

    .card {
        background-color: #FFFFFF;
        border: 1px solid #E1E4E8;
        border-radius: 12px;
        padding: 18px;
        margin-bottom: 14px;
        box-shadow: 0px 1px 2px rgba(0,0,0,0.04);
    }

    .card-important {
        background-color: #FFF8E6;
        border: 1px solid #D6B656;
        border-radius: 12px;
        padding: 18px;
        margin-bottom: 14px;
    }

    .seccion {
        font-size: 19px;
        font-weight: 800;
        color: #003865;
        margin-top: 0px;
        margin-bottom: 12px;
    }

    .mensaje-ok {
        background-color: #EAF7EA;
        border-left: 5px solid #2E7D32;
        padding: 14px 16px;
        border-radius: 8px;
        color: #1B5E20;
        font-size: 15px;
        margin-top: 18px;
        margin-bottom: 18px;
    }

    .mensaje-info {
        background-color: #F4F7FA;
        border-left: 5px solid #003865;
        padding: 14px 16px;
        border-radius: 8px;
        color: #333333;
        font-size: 15px;
        margin-top: 18px;
        margin-bottom: 18px;
    }

    .footer-app {
        text-align: center;
        color: #888888;
        font-size: 12px;
        margin-top: 34px;
    }

    div.stButton > button:first-child {
        width: 100%;
        background-color: #003865;
        color: white;
        font-weight: 700;
        border-radius: 8px;
        height: 3rem;
        border: none;
    }

    div.stButton > button:first-child:hover {
        background-color: #005A8B;
        color: white;
        border: none;
    }

    div.stDownloadButton > button:first-child {
        width: 100%;
        background-color: #006B3F;
        color: white;
        font-weight: 700;
        border-radius: 8px;
        height: 3rem;
        border: none;
    }

    div.stDownloadButton > button:first-child:hover {
        background-color: #00804B;
        color: white;
        border: none;
    }

    [data-testid="stDataFrame"] {
        border: 1px solid #E1E4E8;
        border-radius: 8px;
    }
    </style>
    """,
    unsafe_allow_html=True
)


# =====================================================
# FUNCIONES DE LIMPIEZA Y FORMATO
# =====================================================

def mostrar_logo_centrado(ruta_logo: Path, ancho_px: int = 112) -> None:
    """
    Muestra el logo centrado usando HTML/base64.
    """

    if not ruta_logo.exists():
        return

    with open(ruta_logo, "rb") as f:
        logo_base64 = base64.b64encode(f.read()).decode()

    st.markdown(
        f"""
        <div class="logo-container">
            <img src="data:image/png;base64,{logo_base64}" style="max-width:{ancho_px}px;">
        </div>
        """,
        unsafe_allow_html=True
    )


def limpiar_texto(valor) -> str:
    """
    Limpia valores de texto.
    """

    if pd.isna(valor):
        return ""

    texto = str(valor).strip()

    if texto.lower() in ["nan", "none", "nat"]:
        return ""

    texto = re.sub(r"\s+", " ", texto)

    return texto


def formatear_cantidad(valor, max_decimales: int = 4) -> str:
    """
    Formatea cantidades sin ceros innecesarios.
    Ejemplos:
    18.380000 -> 18.38
    35.100000 -> 35.1
    10.000000 -> 10
    """

    try:
        numero = float(valor)
    except Exception:
        return ""

    if pd.isna(numero):
        return ""

    texto = f"{numero:,.{max_decimales}f}"
    texto = texto.rstrip("0").rstrip(".")

    if texto == "-0":
        texto = "0"

    return texto


def es_codigo_logistico(codigo: str) -> bool:
    """
    Detecta códigos logísticos que no son el código principal del producto.
    """

    codigo = limpiar_texto(codigo)

    if codigo == "":
        return False

    if re.fullmatch(r"20\d{8,}\s*(SA|LF)?", codigo, flags=re.IGNORECASE):
        return True

    if re.fullmatch(r"F\d{5,}", codigo, flags=re.IGNORECASE):
        return True

    if re.search(r"\b20\d{8,}\b", codigo) and len(codigo) > 8:
        return True

    return False


def extraer_codigo_desde_materia(materia_original: str) -> str:
    """
    Extrae código cuando la columna Código viene vacía o viene con un código logístico.

    Prioridad:
    1. Primer código útil dentro de paréntesis, ejemplo 17B71ASK, 17286ASP, 35009*60*01A.
    2. Código E si no existe otro código útil, ejemplo E10073.
    """

    texto = limpiar_texto(materia_original)

    if texto == "":
        return ""

    # Buscar candidatos dentro de paréntesis
    parentesis = re.findall(r"\(([^)]{3,100})\)", texto)

    for item in parentesis:
        candidato = limpiar_texto(item)

        # Saltar unidades o textos con logística
        if re.search(r"\bKg\b|\bL\b|unidad|unid", candidato, flags=re.IGNORECASE):
            continue

        if re.search(r"\b20\d{8,}\b", candidato):
            continue

        if re.search(r"\bF\d{5,}\b", candidato, flags=re.IGNORECASE):
            continue

        # Si contiene código E y más logística, no usarlo como candidato principal
        if re.search(r"\bE\d{4,}\b", candidato, flags=re.IGNORECASE) and len(candidato.split()) > 1:
            continue

        # Código con asterisco
        if re.search(r"\d+\*\d+", candidato):
            return candidato.upper()

        # Código alfanumérico compacto
        if re.fullmatch(r"[A-Z0-9][A-Z0-9*\-/.]{3,30}", candidato, flags=re.IGNORECASE):
            return candidato.upper()

    # Código con asterisco fuera de paréntesis
    match_ast = re.search(r"\b\d{3,}\*\d+(?:\*\d+)?[A-Z0-9]*\b", texto, flags=re.IGNORECASE)
    if match_ast:
        return match_ast.group(0).upper()

    # Código E como última alternativa
    match_e = re.search(r"\bE\d{4,}\b", texto, flags=re.IGNORECASE)
    if match_e:
        return match_e.group(0).upper()

    return ""


def normalizar_codigo(codigo_original, materia_original: str) -> str:
    """
    Mantiene el Código de la base si viene bien.
    Si viene vacío o logístico, intenta obtenerlo desde la materia prima.
    """

    codigo = limpiar_texto(codigo_original)
    candidato = extraer_codigo_desde_materia(materia_original)

    if codigo == "" and candidato != "":
        return candidato

    if es_codigo_logistico(codigo) and candidato != "":
        return candidato

    return codigo


def limpiar_nombre_materia_prima(materia_original: str) -> str:
    """
    Limpia el nombre visible de la materia prima.

    Ejemplo:
    Acesulfamo de Potasio (17B71ASK) (E56685 2000304212 SA F0000014161)
    -> Acesulfamo de Potasio
    """

    texto = limpiar_texto(materia_original)

    if texto == "":
        return ""

    # Quitar contenido entre paréntesis
    texto = re.sub(r"\([^)]*\)", "", texto)

    # Quitar códigos logísticos que queden fuera
    texto = re.sub(r"\bE\d{4,}\b", "", texto, flags=re.IGNORECASE)
    texto = re.sub(r"\bF\d{5,}\b", "", texto, flags=re.IGNORECASE)
    texto = re.sub(r"\b20\d{8,}\b", "", texto)
    texto = re.sub(r"\bSA\b", "", texto, flags=re.IGNORECASE)
    texto = re.sub(r"\bLF\b", "", texto, flags=re.IGNORECASE)

    # Quitar códigos con asterisco al final
    texto = re.sub(r"\b\d{3,}\*\d+(?:\*\d+)?[A-Z0-9]*\b\s*$", "", texto, flags=re.IGNORECASE)

    texto = re.sub(r"\s+", " ", texto).strip(" -_/.")

    return texto


def nombre_archivo_seguro(texto: str) -> str:
    """
    Limpia texto para usarlo como nombre de archivo.
    """

    reemplazos = {
        "/": "-",
        "\\": "-",
        ":": "-",
        "*": "",
        "?": "",
        '"': "",
        "<": "",
        ">": "",
        "|": ""
    }

    texto = str(texto)

    for malo, bueno in reemplazos.items():
        texto = texto.replace(malo, bueno)

    texto = texto.replace(" ", "_")

    return texto


# =====================================================
# CARGA DE BBDD
# =====================================================

@st.cache_data
def cargar_bbdd() -> pd.DataFrame:
    """
    Carga, limpia y normaliza la BBDD Parquet.
    """

    if not ARCHIVO_BBDD.exists():
        st.error("No se encontró el archivo BBDD_UNIDADES_MATERIAS_PRIMAS.parquet en el repositorio.")
        st.stop()

    df = pd.read_parquet(ARCHIVO_BBDD)

    df.columns = [str(col).strip() for col in df.columns]

    columnas_necesarias = [
        "Tipo de elaboración",
        "Materia prima",
        "Código",
        "Unidad medida",
        "Cantidad por unidad"
    ]

    faltantes = [col for col in columnas_necesarias if col not in df.columns]

    if faltantes:
        st.error("La BBDD no tiene las columnas necesarias: " + ", ".join(faltantes))
        st.stop()

    df = df[columnas_necesarias].copy()

    df["Tipo de elaboración"] = df["Tipo de elaboración"].apply(limpiar_texto)
    df["Materia prima original"] = df["Materia prima"].apply(limpiar_texto)

    df["Código"] = df.apply(
        lambda row: normalizar_codigo(row["Código"], row["Materia prima original"]),
        axis=1
    )

    df["Materia prima"] = df["Materia prima original"].apply(limpiar_nombre_materia_prima)

    df["Unidad medida"] = (
        df["Unidad medida"]
        .fillna("Sin unidad explícita")
        .astype(str)
        .str.strip()
    )

    df["Unidad medida"] = df["Unidad medida"].replace(
        {
            "": "Sin unidad explícita",
            "nan": "Sin unidad explícita",
            "None": "Sin unidad explícita"
        }
    )

    df["Cantidad por unidad"] = pd.to_numeric(df["Cantidad por unidad"], errors="coerce")

    df = df.dropna(
        subset=[
            "Tipo de elaboración",
            "Materia prima",
            "Cantidad por unidad"
        ]
    ).copy()

    df = df[
        (df["Tipo de elaboración"] != "") &
        (df["Materia prima"] != "")
    ].copy()

    # Consolidar solo duplicados exactos posteriores a limpieza
    df = (
        df
        .groupby(
            [
                "Tipo de elaboración",
                "Materia prima",
                "Código",
                "Unidad medida"
            ],
            as_index=False
        )
        .agg(Cantidad_por_unidad=("Cantidad por unidad", "sum"))
    )

    df = df.rename(columns={"Cantidad_por_unidad": "Cantidad por unidad"})

    df = df.sort_values(
        ["Tipo de elaboración", "Materia prima"],
        ascending=[True, True]
    ).reset_index(drop=True)

    return df


# =====================================================
# FUNCIONES DE CÁLCULO
# =====================================================

def generar_calculo(df_bbdd: pd.DataFrame, tipo_elaboracion: str, unidades: int) -> pd.DataFrame:
    """
    Calcula materias primas requeridas.
    """

    df_resultado = df_bbdd[df_bbdd["Tipo de elaboración"] == tipo_elaboracion].copy()

    df_resultado = df_resultado.sort_values("Materia prima").reset_index(drop=True)

    df_resultado["Número de unidades"] = unidades
    df_resultado["Cantidad requerida"] = df_resultado["Cantidad por unidad"] * unidades

    df_resultado = df_resultado[
        [
            "Materia prima",
            "Código",
            "Cantidad por unidad",
            "Número de unidades",
            "Cantidad requerida",
            "Unidad medida"
        ]
    ].copy()

    return df_resultado


def crear_tabla_resultado(df_resultado: pd.DataFrame) -> pd.DataFrame:
    """
    Tabla simple para mostrar.
    """

    df_mostrar = df_resultado[
        [
            "Materia prima",
            "Código",
            "Cantidad requerida",
            "Unidad medida"
        ]
    ].copy()

    df_mostrar["Cantidad requerida"] = df_mostrar["Cantidad requerida"].apply(
        lambda x: formatear_cantidad(x, max_decimales=4)
    )

    return df_mostrar


def crear_tabla_detalle(df_resultado: pd.DataFrame) -> pd.DataFrame:
    """
    Tabla de detalle unitario.
    """

    df_detalle = df_resultado[
        [
            "Materia prima",
            "Código",
            "Cantidad por unidad",
            "Número de unidades",
            "Cantidad requerida",
            "Unidad medida"
        ]
    ].copy()

    df_detalle["Cantidad por unidad"] = df_detalle["Cantidad por unidad"].apply(
        lambda x: formatear_cantidad(x, max_decimales=6)
    )

    df_detalle["Cantidad requerida"] = df_detalle["Cantidad requerida"].apply(
        lambda x: formatear_cantidad(x, max_decimales=4)
    )

    return df_detalle


# =====================================================
# EXPORTACIÓN A EXCEL
# =====================================================

def convertir_excel(
    df_resultado: pd.DataFrame,
    tipo_elaboracion: str,
    fecha_calculo: date,
    unidades: int
) -> BytesIO:
    """
    Genera Excel descargable:
    - Resultado
    - Detalle unitario
    """

    output = BytesIO()

    df_simple_excel = df_resultado[
        [
            "Materia prima",
            "Código",
            "Cantidad requerida",
            "Unidad medida"
        ]
    ].copy()

    df_detalle_excel = df_resultado[
        [
            "Materia prima",
            "Código",
            "Cantidad por unidad",
            "Número de unidades",
            "Cantidad requerida",
            "Unidad medida"
        ]
    ].copy()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_simple_excel.to_excel(writer, index=False, sheet_name="Resultado")
        df_detalle_excel.to_excel(writer, index=False, sheet_name="Detalle unitario")

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        fill_header = PatternFill("solid", fgColor="1F4E78")
        font_header = Font(color="FFFFFF", bold=True)

        fill_info = PatternFill("solid", fgColor="EAF3F8")
        font_info = Font(color="003865", bold=True)

        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]

            worksheet.insert_rows(1, 4)

            worksheet["A1"] = "Calculadora de Unidades - Materias Primas"
            worksheet["A1"].font = Font(bold=True, size=14, color="003865")

            worksheet["A2"] = f"Tipo de elaboración: {tipo_elaboracion}"
            worksheet["A3"] = f"Número de unidades: {unidades}"
            worksheet["A4"] = f"Fecha cálculo: {fecha_calculo.strftime('%d-%m-%Y')}"

            for row in range(1, 5):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.fill = fill_info
                    cell.font = font_info
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

            for cell in worksheet[5]:
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            for row in worksheet.iter_rows(min_row=6):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0.####"

            worksheet.freeze_panes = "A6"
            worksheet.auto_filter.ref = worksheet.dimensions

            for col in worksheet.columns:
                col_letter = col[0].column_letter
                max_len = 0

                for cell in col:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(value))

                worksheet.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

    output.seek(0)
    return output



# =====================================================
# ACCESO, SECRETS Y GOOGLE SHEETS
# =====================================================

def obtener_secret(ruta: tuple[str, ...], obligatorio: bool = True, defecto=""):
    """Obtiene un valor de st.secrets usando una ruta segura."""
    actual = st.secrets
    try:
        for clave in ruta:
            actual = actual[clave]
        return actual
    except Exception:
        if obligatorio:
            st.error("Falta configurar el secreto: " + ".".join(ruta))
            st.stop()
        return defecto


def verificar_acceso() -> None:
    """Protege la aplicación con una clave almacenada en st.secrets."""
    if st.session_state.get("acceso_autorizado", False):
        return

    mostrar_logo_centrado(ARCHIVO_LOGO, ancho_px=112)
    st.markdown('<div class="titulo-principal">Acceso a la aplicación</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtitulo">Proyecto CCU - Elaboración</div>', unsafe_allow_html=True)

    with st.form("form_acceso", clear_on_submit=False):
        clave_ingresada = st.text_input("Clave de acceso", type="password")
        ingresar = st.form_submit_button("Ingresar")

    if ingresar:
        clave_configurada = str(obtener_secret(("app", "clave_acceso")))
        hash_ingresado = hashlib.sha256(clave_ingresada.encode("utf-8")).digest()
        hash_configurado = hashlib.sha256(clave_configurada.encode("utf-8")).digest()

        if hash_ingresado == hash_configurado:
            st.session_state["acceso_autorizado"] = True
            st.rerun()
        else:
            st.error("Clave incorrecta.")

    st.stop()


def valor_fecha_iso(valor) -> str:
    """Convierte date, datetime o texto de fecha al formato ISO."""
    if valor is None or pd.isna(valor):
        return ""
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    fecha = pd.to_datetime(valor, errors="coerce", dayfirst=True)
    return "" if pd.isna(fecha) else fecha.date().isoformat()


def crear_tabla_registro(df_resultado: pd.DataFrame) -> pd.DataFrame:
    """Crea la tabla editable de trazabilidad para las materias primas."""
    df = df_resultado[[
        "Materia prima", "Código", "Cantidad requerida", "Unidad medida"
    ]].copy()
    df["Proveedor"] = ""
    df["Lote"] = ""
    df["Fecha elaboración"] = pd.NaT
    df["Fecha vencimiento"] = pd.NaT
    df["N° bolsas / bidones"] = ""
    df["N° bidón / bolsa"] = ""
    df["Inspección visual"] = "Cumple"
    df["Observación"] = ""
    return df


def validar_registro(orden: str, destino: str, operador: str, df: pd.DataFrame) -> list[str]:
    """Valida encabezado, trazabilidad, fechas e inspección visual."""
    errores: list[str] = []

    if not orden.strip():
        errores.append("Debe ingresar la orden de elaboración.")
    if not destino.strip():
        errores.append("Debe ingresar el destino.")
    if not operador.strip():
        errores.append("Debe ingresar el operador.")

    obligatorias = [
        "Proveedor", "Lote", "Fecha vencimiento",
        "N° bolsas / bidones", "N° bidón / bolsa", "Inspección visual"
    ]

    for columna in obligatorias:
        if columna not in df.columns:
            errores.append(f"Falta la columna obligatoria: {columna}.")
            continue
        serie = df[columna]
        vacios = serie.isna() | serie.astype(str).str.strip().str.lower().isin(["", "nat", "nan", "none"])
        if vacios.any():
            errores.append(f"Complete todos los valores de: {columna}.")

    fechas_vencimiento = pd.to_datetime(df.get("Fecha vencimiento"), errors="coerce")
    if fechas_vencimiento.isna().any():
        errores.append("Existe una fecha de vencimiento inválida.")
    elif (fechas_vencimiento.dt.date < date.today()).any():
        errores.append("Existe al menos una materia prima vencida.")

    fechas_elaboracion = pd.to_datetime(df.get("Fecha elaboración"), errors="coerce")
    validas = fechas_elaboracion.notna() & fechas_vencimiento.notna()
    if validas.any() and (fechas_elaboracion[validas] > fechas_vencimiento[validas]).any():
        errores.append("Una fecha de elaboración es posterior a su vencimiento.")

    inspeccion = df.get("Inspección visual", pd.Series(dtype=str)).astype(str).str.strip().str.lower()
    if inspeccion.eq("no cumple").any():
        errores.append("No se puede guardar con una inspección visual 'No cumple'.")

    return list(dict.fromkeys(errores))


def construir_payload(
    fecha_calculo: date,
    orden: str,
    tipo_elaboracion: str,
    unidades: int,
    destino: str,
    turno: str,
    operador: str,
    observacion_general: str,
    df_registro: pd.DataFrame,
) -> dict:
    """Construye el JSON enviado al Apps Script."""
    id_registro = str(uuid.uuid4())
    detalle = []

    for _, fila in df_registro.iterrows():
        detalle.append({
            "id_registro": id_registro,
            "materia_prima": limpiar_texto(fila.get("Materia prima")),
            "codigo": limpiar_texto(fila.get("Código")),
            "cantidad_requerida": float(fila.get("Cantidad requerida", 0)),
            "unidad_medida": limpiar_texto(fila.get("Unidad medida")),
            "proveedor": limpiar_texto(fila.get("Proveedor")),
            "lote": limpiar_texto(fila.get("Lote")),
            "fecha_elaboracion": valor_fecha_iso(fila.get("Fecha elaboración")),
            "fecha_vencimiento": valor_fecha_iso(fila.get("Fecha vencimiento")),
            "numero_envases": limpiar_texto(fila.get("N° bolsas / bidones")),
            "numero_trazabilidad": limpiar_texto(fila.get("N° bidón / bolsa")),
            "inspeccion_visual": limpiar_texto(fila.get("Inspección visual")),
            "observacion": limpiar_texto(fila.get("Observación")),
        })

    return {
        "token": str(obtener_secret(("google_sheets", "token"))),
        "accion": "guardar_elaboracion",
        "elaboracion": {
            "id_registro": id_registro,
            "fecha": fecha_calculo.isoformat(),
            "orden_elaboracion": orden.strip(),
            "tipo_elaboracion": tipo_elaboracion,
            "unidades": int(unidades),
            "destino": destino.strip(),
            "turno": turno,
            "operador": operador.strip(),
            "observacion_general": observacion_general.strip(),
            "fecha_registro": datetime.now().isoformat(timespec="seconds"),
        },
        "detalle": detalle,
    }


def enviar_a_google_sheets(payload: dict) -> dict:
    """Envía un registro JSON al Apps Script configurado en secrets."""
    url = str(obtener_secret(("google_sheets", "apps_script_url")))
    if not url.startswith("https://script.google.com/"):
        raise ValueError("La URL de Apps Script configurada no es válida.")

    solicitud = Request(
        url=url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json; charset=utf-8"},
        method="POST",
    )

    try:
        with urlopen(solicitud, timeout=30) as respuesta:
            contenido = respuesta.read().decode("utf-8")
    except HTTPError as error:
        detalle = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Google respondió HTTP {error.code}: {detalle[:300]}") from error
    except URLError as error:
        raise RuntimeError(f"No fue posible conectar con Google Sheets: {error.reason}") from error

    try:
        resultado = json.loads(contenido)
    except json.JSONDecodeError as error:
        raise RuntimeError("Apps Script devolvió una respuesta no válida.") from error

    if not resultado.get("ok", False):
        raise RuntimeError(resultado.get("mensaje", "El registro fue rechazado por Apps Script."))

    return resultado


# =====================================================
# INICIO Y CARGA DE DATOS
# =====================================================

verificar_acceso()
df_bbdd = cargar_bbdd()


# =====================================================
# ENCABEZADO
# =====================================================

mostrar_logo_centrado(ARCHIVO_LOGO, ancho_px=112)

st.markdown(
    """
    <div class="titulo-principal">Calculadora de Materias Primas</div>
    <div class="subtitulo">Proyecto CCU - Elaboración | Sala de unidades</div>
    """,
    unsafe_allow_html=True
)

with st.sidebar:
    st.caption("Sesión autorizada")
    if st.button("Cerrar sesión"):
        st.session_state.clear()
        st.rerun()


# =====================================================
# FORMULARIO PRINCIPAL
# =====================================================

st.markdown('<div class="card-important">', unsafe_allow_html=True)
st.markdown('<div class="seccion">Datos del cálculo</div>', unsafe_allow_html=True)

with st.form("form_calculadora"):
    col1, col2, col3 = st.columns([1.1, 2.6, 1.0])

    with col1:
        fecha_calculo = st.date_input("Fecha", value=date.today())

    with col2:
        tipos_elaboracion = sorted(df_bbdd["Tipo de elaboración"].unique())
        tipo_seleccionado = st.selectbox("Tipo de elaboración", options=tipos_elaboracion)

    with col3:
        unidades = st.number_input("Número de unidades", min_value=1, value=1, step=1)

    calcular = st.form_submit_button("Calcular materias primas")

st.markdown('</div>', unsafe_allow_html=True)

if calcular:
    st.session_state["df_resultado"] = generar_calculo(df_bbdd, tipo_seleccionado, unidades)
    st.session_state["fecha_calculo"] = fecha_calculo
    st.session_state["tipo_seleccionado"] = tipo_seleccionado
    st.session_state["unidades"] = unidades
    st.session_state["df_registro"] = crear_tabla_registro(st.session_state["df_resultado"])


# =====================================================
# RESULTADO Y REGISTRO
# =====================================================

if "df_resultado" in st.session_state:
    df_resultado = st.session_state["df_resultado"]
    fecha_calculo = st.session_state["fecha_calculo"]
    tipo_seleccionado = st.session_state["tipo_seleccionado"]
    unidades = st.session_state["unidades"]

    df_mostrar = crear_tabla_resultado(df_resultado)
    df_detalle = crear_tabla_detalle(df_resultado)

    st.markdown(
        f"""
        <div class="mensaje-ok">
            <b>Cálculo generado correctamente.</b><br>
            Elaboración seleccionada: <b>{tipo_seleccionado}</b><br>
            Unidades solicitadas: <b>{unidades}</b><br>
            Materias primas requeridas: <b>{len(df_mostrar)}</b>
        </div>
        """,
        unsafe_allow_html=True
    )

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="seccion">Materias primas requeridas</div>', unsafe_allow_html=True)
    st.dataframe(df_mostrar, use_container_width=True, hide_index=True)

    with st.expander("Ver detalle de cantidad unitaria"):
        st.dataframe(df_detalle, use_container_width=True, hide_index=True)

    st.markdown('</div>', unsafe_allow_html=True)

    archivo_excel = convertir_excel(df_resultado, tipo_seleccionado, fecha_calculo, unidades)
    nombre_descarga = (
        "CALCULO_MP_" + nombre_archivo_seguro(tipo_seleccionado) + "_" +
        fecha_calculo.strftime("%Y%m%d") + ".xlsx"
    )
    st.download_button(
        "Descargar resultado en Excel",
        data=archivo_excel,
        file_name=nombre_descarga,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="seccion">Registro de elaboración y trazabilidad</div>', unsafe_allow_html=True)

    with st.form("form_registro", clear_on_submit=False):
        c1, c2, c3, c4 = st.columns([1.4, 1.3, 0.7, 1.3])
        with c1:
            orden = st.text_input("Orden de elaboración")
        with c2:
            destino = st.text_input("Destino", placeholder="Ejemplo: Jarabe 1")
        with c3:
            turno = st.selectbox("Turno", ["A", "B", "C"])
        with c4:
            operador = st.text_input("Operador")

        observacion_general = st.text_area("Observación general", height=80)

        df_registro_editado = st.data_editor(
            st.session_state["df_registro"],
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            disabled=["Materia prima", "Código", "Cantidad requerida", "Unidad medida"],
            column_config={
                "Cantidad requerida": st.column_config.NumberColumn(format="%.4f"),
                "Fecha elaboración": st.column_config.DateColumn(format="DD-MM-YYYY"),
                "Fecha vencimiento": st.column_config.DateColumn(format="DD-MM-YYYY", required=True),
                "Inspección visual": st.column_config.SelectboxColumn(
                    options=["Cumple", "No cumple"], required=True
                ),
            },
            key="editor_trazabilidad",
        )

        guardar = st.form_submit_button("Guardar registro en Google Sheets")

    if guardar:
        st.session_state["df_registro"] = df_registro_editado
        errores = validar_registro(orden, destino, operador, df_registro_editado)

        if errores:
            for error in errores:
                st.error(error)
        else:
            try:
                payload = construir_payload(
                    fecha_calculo, orden, tipo_seleccionado, unidades,
                    destino, turno, operador, observacion_general,
                    df_registro_editado,
                )
                with st.spinner("Guardando registro..."):
                    respuesta = enviar_a_google_sheets(payload)
                st.success(
                    "Registro guardado correctamente. ID: " +
                    respuesta.get("id_registro", payload["elaboracion"]["id_registro"])
                )
            except Exception as error:
                st.error(f"No fue posible guardar el registro: {error}")

    st.markdown('</div>', unsafe_allow_html=True)
else:
    st.markdown(
        """
        <div class="mensaje-info">
            Completa los datos del cálculo y presiona <b>Calcular materias primas</b>.
        </div>
        """,
        unsafe_allow_html=True
    )


# =====================================================
# PIE DE PÁGINA
# =====================================================

st.markdown(
    """
    <div class="footer-app">
        Proyecto CCU - Elaboración | BBDD Parquet + Google Sheets
    </div>
    """,
    unsafe_allow_html=True
)
